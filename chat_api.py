"""
Chat API - Endpoints for chat-first automation
"""
import json
import asyncio
import csv
import io
from datetime import datetime
from typing import Optional, List, Dict, Any
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlmodel import Session, select

from database import get_sync_session, sync_engine
from chat_models import (
    ChatMessage, AutomationRule, RuleExecution, 
    Incident, KnownNode, RuleStatus, IncidentSeverity,
    ChatSession  # Added ChatSession
)
from chat_engine import (
    parse_user_intent, generate_execution_narrative,
    resolve_node, generate_rule_card_text, analyze_output_with_llm
)
from executor import execute_command_on_server
from crypto import decrypt_credential

# Create tables
from sqlmodel import SQLModel
SQLModel.metadata.create_all(sync_engine)

router = APIRouter(prefix="/chat", tags=["Chat"])


# --- Request/Response Models ---
class ChatRequest(BaseModel):
    message: str
    session_id: Optional[int] = None


class CreateSessionRequest(BaseModel):
    title: Optional[str] = "New Session"


class ChatResponse(BaseModel):
    type: str  # "rule" | "clarification" | "chat" | "error"
    message: Optional[str] = None
    rule: Optional[Dict[str, Any]] = None
    rule_id: Optional[int] = None
    clarification: Optional[str] = None
    session_id: Optional[int] = None


class ExecuteRequest(BaseModel):
    rule_id: int


# --- Helper Functions ---
def get_known_nodes() -> Dict[str, Dict]:
    """Get all known nodes as a dict for quick lookup"""
    with get_sync_session() as session:
        nodes = session.exec(select(KnownNode)).all()
        return {
            node.name: {
                "name": node.name,
                "hostname": node.hostname,
                "username": node.username,
                "port": node.port,
                "auth_type": node.auth_type,
                "credential": decrypt_credential(node.credential_encrypted) if node.credential_encrypted else None
            }
            for node in nodes
        }


def save_chat_message(role: str, content: str, rule_id: int = None, session_id: int = None):
    """Save a chat message to history"""
    with get_sync_session() as session:
        msg = ChatMessage(role=role, content=content, rule_id=rule_id, session_id=session_id)
        session.add(msg)
        session.commit()


# --- Endpoints ---

@router.post("/sessions", response_model=dict)
async def create_session(request: CreateSessionRequest):
    """Create a new chat session (thread)"""
    with get_sync_session() as session:
        chat_session = ChatSession(title=request.title)
        session.add(chat_session)
        session.commit()
        session.refresh(chat_session)
        return {"id": chat_session.id, "title": chat_session.title, "created_at": chat_session.created_at}


@router.get("/sessions", response_model=List[dict])
async def list_sessions():
    """List all chat sessions"""
    with get_sync_session() as session:
        sessions = session.exec(select(ChatSession).order_by(ChatSession.created_at.desc())).all()
        return [
            {"id": s.id, "title": s.title, "created_at": s.created_at}
            for s in sessions
        ]


@router.get("/sessions/{session_id}/messages")
async def get_session_messages(session_id: int):
    """Get messages for a specific session"""
    with get_sync_session() as session:
        messages = session.exec(
            select(ChatMessage)
            .where(ChatMessage.session_id == session_id)
            .order_by(ChatMessage.created_at.asc())
        ).all()
        
        result = []
        for msg in messages:
            rule_data = None
            if msg.rule_id:
                rule = session.get(AutomationRule, msg.rule_id)
                if rule:
                    rule_data = {
                        "name": rule.name,
                        "nodes": json.loads(rule.nodes_json),
                        "commands": json.loads(rule.commands_json) if rule.commands_json else [],
                        "execution_mode": rule.execution_mode,
                         # Legacy map for fallback
                        "command": rule.command,
                    }
            
            result.append({
                "id": msg.id,
                "role": msg.role,
                "content": msg.content,
                "rule_id": msg.rule_id,
                "rule": rule_data,
                "created_at": msg.created_at.isoformat()
            })
        return result


@router.delete("/sessions/{session_id}")
async def delete_session(session_id: int):
    """Delete a chat session and its associated messages and rules"""
    with get_sync_session() as session:
        # First delete associated messages
        messages = session.exec(
            select(ChatMessage).where(ChatMessage.session_id == session_id)
        ).all()
        for msg in messages:
            session.delete(msg)
        
        # Delete associated rules
        rules = session.exec(
            select(AutomationRule).where(AutomationRule.session_id == session_id)
        ).all()
        for rule in rules:
            session.delete(rule)
        
        # Delete the session itself
        chat_session = session.get(ChatSession, session_id)
        if chat_session:
            session.delete(chat_session)
            session.commit()
            return {"status": "deleted", "session_id": session_id}
        else:
            raise HTTPException(status_code=404, detail="Session not found")


@router.post("/upload-nodes")
async def upload_nodes_file(file: UploadFile = File(...)):
    """
    Upload a CSV or Excel file containing node definitions.
    Expected columns: hostname, username, auth_type (optional), port (optional)
    Returns a summary and parsed nodes for use in rule creation.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    # Check file extension
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
    
    content = await file.read()
    nodes = []
    
    try:
        if filename.endswith('.csv'):
            # Parse CSV
            text = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                node = {
                    "hostname": row.get('hostname', row.get('ip', row.get('host', ''))).strip(),
                    "name": row.get('name', row.get('hostname', row.get('ip', ''))).strip(),
                    "auth": {
                        "username": row.get('username', row.get('user', 'ubuntu')).strip(),
                        "type": row.get('auth_type', 'key').strip()
                    }
                }
                if node["hostname"]:
                    nodes.append(node)
        else:
            # Parse Excel - requires openpyxl
            try:
                import openpyxl
                from io import BytesIO
                wb = openpyxl.load_workbook(BytesIO(content))
                sheet = wb.active
                headers = [cell.value.lower() if cell.value else '' for cell in sheet[1]]
                
                hostname_col = headers.index('hostname') if 'hostname' in headers else (
                    headers.index('ip') if 'ip' in headers else headers.index('host') if 'host' in headers else -1
                )
                username_col = headers.index('username') if 'username' in headers else (
                    headers.index('user') if 'user' in headers else -1
                )
                name_col = headers.index('name') if 'name' in headers else -1
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    hostname = str(row[hostname_col]).strip() if hostname_col >= 0 and row[hostname_col] else ''
                    username = str(row[username_col]).strip() if username_col >= 0 and row[username_col] else 'ubuntu'
                    name = str(row[name_col]).strip() if name_col >= 0 and row[name_col] else hostname
                    
                    if hostname:
                        nodes.append({
                            "hostname": hostname,
                            "name": name or hostname,
                            "auth": {"username": username, "type": "key"}
                        })
            except ImportError:
                raise HTTPException(status_code=500, detail="Excel parsing requires openpyxl. Install with: pip install openpyxl")
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    
    if not nodes:
        raise HTTPException(status_code=400, detail="No valid nodes found in file")
    
    # Generate summary text for chat input
    summary = f"Imported {len(nodes)} nodes from {file.filename}"
    preview = nodes[:5]  # First 5 for preview
    
    return {
        "node_count": len(nodes),
        "nodes": nodes,
        "preview": preview,
        "summary": summary,
        "suggested_prompt": f"Run health check on all {len(nodes)} imported nodes"
    }


@router.get("/report/{rule_id}")
async def get_report(rule_id: int, format: str = "json"):
    """
    Generate and return a report for a specific rule execution.
    Supports formats: json, csv, excel
    Format: Nodename, Nodetype, Nodeip, Command, Deviationstatus, Deviationremarks, Executionstatus, Executionremarks, Parserresult
    """
    with get_sync_session() as session:
        # Get the rule
        rule = session.get(AutomationRule, rule_id)
        if not rule:
            raise HTTPException(status_code=404, detail="Rule not found")
        
        # Get the latest execution for this rule
        execution = session.exec(
            select(RuleExecution)
            .where(RuleExecution.rule_id == rule_id)
            .order_by(RuleExecution.started_at.desc())
        ).first()
        
        # Parse rule data
        nodes = json.loads(rule.nodes_json) if rule.nodes_json else []
        commands = json.loads(rule.commands_json) if rule.commands_json else []
        timeline = json.loads(execution.timeline_json) if execution and execution.timeline_json else []
        
        # Build tabular report rows by tracking command context in timeline
        # Each row: Nodename, Nodetype, Nodeip, Command, Deviationstatus, Deviationremarks, Executionstatus, Executionremarks, Parserresult
        report_rows = []
        
        # Parse timeline to extract results per command
        # Timeline structure: "Running command: X" -> "Output received" -> "Analyzing..." -> "✓ All clear" or "⚠️ Condition met"
        command_results = {}
        current_cmd = None
        
        for event in timeline:
            event_msg = event.get('message', '')
            event_status = event.get('status', '')
            event_type = event.get('type', '')
            
            # Detect command start
            if 'Running command:' in event_msg:
                current_cmd = event_msg.replace('Running command:', '').strip()
                command_results[current_cmd] = {
                    "executed": True,
                    "status": "OK",
                    "remarks": "-"
                }
            
            # Detect condition met (NOK case)
            elif current_cmd and ('⚠️' in event_msg or 'Condition met' in event_msg or event_status == 'warning'):
                command_results[current_cmd]["status"] = "NOK"
                # Extract the remarks after "Condition met:"
                if 'Condition met:' in event_msg:
                    remarks = event_msg.split('Condition met:')[-1].strip()
                else:
                    remarks = event_msg.replace('⚠️', '').strip()
                command_results[current_cmd]["remarks"] = remarks
            
            # Detect success (OK case)
            elif current_cmd and ('✓' in event_msg or 'All clear' in event_msg):
                if command_results[current_cmd]["status"] != "NOK":  # Don't override NOK
                    command_results[current_cmd]["status"] = "OK"
                    remarks = event_msg.replace('✓', '').replace('All clear:', '').strip()
                    command_results[current_cmd]["remarks"] = remarks if remarks else "Success"
            
            # Detect incident type events
            elif event_type == 'incident':
                incident_msg = event.get('details', {}).get('summary', event_msg)
                # Try to associate with current command
                if current_cmd:
                    command_results[current_cmd]["status"] = "NOK"
                    command_results[current_cmd]["remarks"] = incident_msg[:200]
        
        # Now build report rows for each node + command combination
        for node in nodes:
            node_name = node.get('name', 'Unknown')
            node_type = node.get('auth', {}).get('type', 'SSH').upper()
            node_ip = node.get('hostname', 'N/A')
            
            for cmd in commands:
                cmd_text = cmd.get('cmd', 'N/A')
                cmd_logic = cmd.get('logic', '')
                
                # Find matching result from timeline
                result = command_results.get(cmd_text, {})
                
                deviation_status = result.get("status", "OK")
                deviation_remarks = result.get("remarks", "-")
                execution_status = "Executed" if result.get("executed") else "-"
                parser_result = f"Logic: {cmd_logic}" if cmd_logic and deviation_status == "NOK" else "-"
                
                report_rows.append({
                    "Nodename": node_name,
                    "Nodetype": node_type,
                    "Nodeip": node_ip,
                    "Command": cmd_text,
                    "Deviationstatus": deviation_status,
                    "Deviationremarks": deviation_remarks,
                    "Executionstatus": execution_status,
                    "Executionremarks": "-",
                    "Parserresult": parser_result
                })
        
        # Build full report data
        report_data = {
            "rule_name": rule.name,
            "rule_id": rule_id,
            "execution_id": execution.id if execution else None,
            "status": execution.status.value if execution else "not_executed",
            "started_at": execution.started_at.isoformat() if execution else None,
            "completed_at": execution.completed_at.isoformat() if execution and execution.completed_at else None,
            "generated_at": datetime.utcnow().isoformat(),
            "rows": report_rows
        }
        
        if format == "json":
            return report_data
        
        elif format == "csv":
            # Generate CSV with tabular format
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Header row
            writer.writerow([
                "Nodename", "Nodetype", "Nodeip", "Command", 
                "Deviationstatus", "Deviationremarks", 
                "Executionstatus", "Executionremarks", "Parserresult"
            ])
            
            # Data rows
            for row in report_rows:
                writer.writerow([
                    row["Nodename"], row["Nodetype"], row["Nodeip"], row["Command"],
                    row["Deviationstatus"], row["Deviationremarks"],
                    row["Executionstatus"], row["Executionremarks"], row["Parserresult"]
                ])
            
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=report_rule_{rule_id}.csv"}
            )
        
        elif format == "excel":
            # Generate Excel with tabular format
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Execution Report"
                
                # Styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                nok_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Title
                ws['A1'] = f"Athena Agent - Execution Report: {rule.name}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:I1')
                
                ws['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
                ws['A2'].font = Font(italic=True, size=10)
                
                # Header row
                headers = [
                    "Nodename", "Nodetype", "Nodeip", "Command", 
                    "Deviationstatus", "Deviationremarks", 
                    "Executionstatus", "Executionremarks", "Parserresult"
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                
                # Data rows
                for row_idx, row_data in enumerate(report_rows, 5):
                    for col_idx, key in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=row_data[key])
                        cell.border = thin_border
                        
                        # Color code status columns
                        if key == "Deviationstatus":
                            if row_data[key] == "OK":
                                cell.fill = ok_fill
                            elif row_data[key] == "NOK":
                                cell.fill = nok_fill
                
                # Adjust column widths
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 25
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 50
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 20
                ws.column_dimensions['I'].width = 20
                
                # Save to bytes
                from io import BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                return StreamingResponse(
                    iter([output.getvalue()]),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=report_rule_{rule_id}.xlsx"}
                )
            except ImportError:
                raise HTTPException(status_code=500, detail="Excel export requires openpyxl. Install with: pip install openpyxl")
        
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use: json, csv, or excel")


@router.post("/message", response_model=ChatResponse)
async def send_message(request: ChatRequest):
    """
    Main chat endpoint - accepts natural language, returns rule or response
    """
    session_id = request.session_id
    
    # If no session provided, Create one? Or Error?
    # Logic: If FE sends no session, we create one implicit or explicit? 
    # Better to assume FE manages sessions. But for fallback:
    if not session_id:
        # Create a temp/default session
         with get_sync_session() as session:
            new_session = ChatSession(title="Ad-hoc Session")
            session.add(new_session)
            session.commit()
            session_id = new_session.id
    
    # Save user message
    save_chat_message("user", request.message, session_id=session_id)
    
    # Get known nodes for context
    known_nodes = get_known_nodes()
    known_node_names = list(known_nodes.keys())
    
    # Parse user intent with LLM
    result = parse_user_intent(request.message, known_node_names)
    
    if result.get("type") == "clarification":
        # Need more info
        clarification = result.get("clarification", "Could you provide more details?")
        save_chat_message("assistant", clarification, session_id=session_id)
        return ChatResponse(
            type="clarification",
            clarification=clarification,
            session_id=session_id
        )
    
    elif result.get("type") == "rule":
        # Got a valid rule
        rule_data = result.get("rule", {})
        
        # Try to resolve nodes
        resolved_nodes = []
        unresolved_nodes = []
        
        for node_item in rule_data.get("nodes", []):
            # Handler if LLM returns string (name) or dict (config)
            if isinstance(node_item, str):
                node_name = node_item
                overrides = {}
            else:
                node_name = node_item.get("name") or node_item.get("hostname")
                overrides = node_item
            
            resolved, config = resolve_node(node_name, known_nodes)
            
            if resolved:
                # Merge overrides from LLM (e.g. specific auth provided in chat)
                if isinstance(node_item, dict):
                    config.update(node_item)
                resolved_nodes.append(config)
            else:
                # If valid config provided by LLM (has hostname), accept it even if not in known_nodes
                if isinstance(node_item, dict) and node_item.get("hostname"):
                     resolved_nodes.append(node_item)
                else:
                     unresolved_nodes.append(node_name)
        
        # If we have unresolved nodes, ask for clarification
        if unresolved_nodes and not resolved_nodes:
            # Check if we have at least one valid node
            pass # Logic handled above covers partial resolution? 
            # If strictly NO resolved nodes and ANY unresolved:
            clarification = f"I don't recognize '{unresolved_nodes[0]}'. What's its IP address? (e.g., 192.168.1.10)"
            save_chat_message("assistant", clarification, session_id=session_id)
            return ChatResponse(
                type="clarification",
                clarification=clarification,
                session_id=session_id
            )
        
        # Create the rule in database
        with get_sync_session() as session:
            rule = AutomationRule(
                name=rule_data.get("name", "Unnamed Rule"),
                session_id=session_id,
                nodes_json=json.dumps(resolved_nodes),
                commands_json=json.dumps(rule_data.get("commands", [])),
                execution_mode=rule_data.get("execution", "sequential"),
                report_config_json=json.dumps(rule_data.get("report", {})),
                command="DEPRECATED", # Kept for schema compatibility
                status=RuleStatus.CREATED
            )
            session.add(rule)
            session.commit()
            session.refresh(rule)
            rule_id = rule.id
        
        # Generate rule card text
        rule_card = generate_rule_card_text(rule_data)
        save_chat_message("assistant", rule_card, rule_id=rule_id, session_id=session_id)
        
        return ChatResponse(
            type="rule",
            rule=rule_data,
            rule_id=rule_id,
            message=rule_card,
            session_id=session_id
        )
    
    else:
        # General chat response
        message = result.get("message", "I'm here to help automate your infrastructure.")
        save_chat_message("assistant", message, session_id=session_id)
        return ChatResponse(
            type="chat",
            message=message,
            session_id=session_id
        )


@router.post("/execute/{rule_id}")
async def execute_rule(rule_id: int):
    """
    Execute a rule and return execution ID for streaming
    """
    with get_sync_session() as session:
        rule = session.get(AutomationRule, rule_id)
        if not rule:
            raise HTTPException(status_code=404, detail="Rule not found")
        
        # Create execution record
        execution = RuleExecution(
            rule_id=rule_id,
            status=RuleStatus.EXECUTING,
            timeline_json="[]"
        )
        session.add(execution)
        session.commit()
        session.refresh(execution)
        
        return {"execution_id": execution.id, "status": "started"}


@router.get("/execute/{execution_id}/stream")
async def stream_execution(execution_id: int):
    """
    Stream execution timeline events via Server-Sent Events.
    Supports parallel/sequential execution across multiple nodes and commands.
    """
    async def event_generator():
        with get_sync_session() as session:
            execution = session.get(RuleExecution, execution_id)
            if not execution:
                yield f"data: {json.dumps({'error': 'Execution not found'})}\n\n"
                return
            
            rule = session.get(AutomationRule, execution.rule_id)
            if not rule:
                yield f"data: {json.dumps({'error': 'Rule not found'})}\n\n"
                return
            
            # Parse execution configuration
            nodes_config = json.loads(rule.nodes_json)
            commands_config = json.loads(rule.commands_json or "[]")
            
            # Fallback for legacy rules with single command (backward compatibility)
            if not commands_config and rule.command:
                commands_config = [{
                    "cmd": rule.command, 
                    "logic": rule.extraction_hint or "Execute command"
                }]
            
            execution_mode = rule.execution_mode or "sequential"
            
            # Shared queue for events from all nodes
            queue = asyncio.Queue()
            
            # Track overall progress
            timeline = []
            
            async def process_node(node_config: Dict, commands: List[Dict]):
                """Worker to process a single node's automation sequence"""
                node_name = node_config.get("name", node_config.get("hostname", "unknown"))
                hostname = node_config.get("hostname")
                username = node_config.get("auth", {}).get("username", "ubuntu")
                
                # Determine credentials
                credential = None
                known_nodes = get_known_nodes()
                
                # 1. Try inline auth
                auth = node_config.get("auth", {})
                if auth.get("type") == "password":
                    credential = auth.get("password")
                # 2. Try known node match
                elif node_name in known_nodes:
                    credential = known_nodes[node_name].get("credential")
                
                try:
                    # Step 1: Connecting
                    await queue.put({"type": "step", "message": generate_execution_narrative('connecting', {'node': node_name})})
                    
                    import asyncssh
                    
                    connect_options = {
                        "host": hostname,
                        "username": username,
                        "known_hosts": None
                    }
                    
                    if credential:
                        # Simple heuristic: if it looks like a private key (starts with -), use ClientKeys
                        if "-----BEGIN" in credential:
                            connect_options["client_keys"] = [asyncssh.import_private_key(credential)]
                        else:
                            connect_options["password"] = credential
                    else:
                        # If no credential specific for this node, try to use local SSH keys (agent behavior)
                        # We explicitly look for common keys + agent_vm_key
                        import os
                        local_keys = []
                        for key_file in ["id_rsa", "id_ed25519", "agent_vm_key"]:
                            key_path = os.path.expanduser(f"~/.ssh/{key_file}")
                            if os.path.exists(key_path):
                                local_keys.append(key_path)
                        
                        if local_keys:
                            connect_options["client_keys"] = local_keys
                            
                    # Connect
                    async with await asyncssh.connect(**connect_options) as conn:
                        # Step 2: Connected
                        await queue.put({"type": "step", "message": generate_execution_narrative('connected', {'node': node_name}), "status": "success"})
                        
                        # Loop through commands
                        for cmd_config in commands:
                            cmd_str = cmd_config.get("cmd")
                            cmd_logic = cmd_config.get("logic", "")
                            
                            # Step 3: Executing
                            await queue.put({"type": "step", "message": generate_execution_narrative('executing', {'command': cmd_str, 'node': node_name})})
                            
                            result = await conn.run(cmd_str, check=False)
                            
                            # Step 4: Output Received
                            await queue.put({"type": "step", "message": generate_execution_narrative('output_received', {'node': node_name}), "status": "success"})
                            
                            # Step 5: Analyzing
                            await queue.put({"type": "step", "message": generate_execution_narrative('analyzing', {'node': node_name})})
                            
                            # Run LLM analysis in thread pool to execute non-blocking
                            loop = asyncio.get_event_loop()
                            analysis = await loop.run_in_executor(
                                None, 
                                analyze_output_with_llm, 
                                result.stdout, 
                                cmd_logic, 
                                rule.condition or "Check for errors"
                            )
                            
                            # Step 6: Result
                            if analysis.get("condition_met"):
                                await queue.put({
                                    "type": "step", 
                                    "message": generate_execution_narrative('condition_met', {'message': analysis.get('summary'), 'node': node_name}), 
                                    "status": "warning"
                                })
                                
                                # Raise incident if needed
                                # Note: Writing to DB from async worker needs care with sessions. 
                                # We'll just emit the event for now and handle DB update in consumer or use a fresh session
                                await queue.put({
                                    "type": "incident",
                                    "message": generate_execution_narrative('incident_raised', {'message': analysis.get('summary')}),
                                    "severity": analysis.get('severity'),
                                    "node": node_name,
                                    "details": analysis
                                })
                            else:
                                await queue.put({
                                    "type": "step", 
                                    "message": generate_execution_narrative('condition_not_met', {'message': analysis.get('summary'), 'node': node_name}), 
                                    "status": "success"
                                })

                except Exception as e:
                    await queue.put({"type": "error", "message": generate_execution_narrative('failed', {'error': str(e), 'node': node_name})})

            # Start processing
            tasks = []
            if execution_mode == "parallel":
                tasks = [asyncio.create_task(process_node(node, commands_config)) for node in nodes_config]
            else:
                # Sequential: We'll wrap in a task that runs them sequentially
                async def sequential_runner():
                    for node in nodes_config:
                        await process_node(node, commands_config)
                tasks = [asyncio.create_task(sequential_runner())]

            # Consumer Loop
            finished_tasks = 0
            while finished_tasks < len(tasks):
                # Wait for next event or task completion
                # We use a timeout to check task status periodically if queue is empty
                try:
                    event = await asyncio.wait_for(queue.get(), timeout=0.5)
                    
                    # Handle Incident DB saving 'out of band' here if strictly necessary, 
                    # but for simplicity we stream it.
                    # Update timeline for persistence
                    if event.get("type") == "step":
                         timeline.append(event)
                         # Consider persisting timeline to DB periodically or at end
                    
                    yield f"data: {json.dumps(event)}\n\n"
                    
                except asyncio.TimeoutError:
                    # Check if tasks are done
                    finished_tasks = sum(1 for t in tasks if t.done())

            # Consuming remaining events
            while not queue.empty():
                event = await queue.get()
                timeline.append(event)
                yield f"data: {json.dumps(event)}\n\n"

            # Finalize
            execution.status = RuleStatus.COMPLETED
            execution.completed_at = datetime.utcnow()
            execution.timeline_json = json.dumps(timeline)
            rule.last_executed_at = datetime.utcnow()
            session.commit()
            
            yield f"data: {json.dumps({'type': 'complete', 'message': generate_execution_narrative('completed')})}\n\n"
    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        }
    )


@router.get("/history")
async def get_chat_history(limit: int = 50):
    """Get chat history"""
    with get_sync_session() as session:
        messages = session.exec(
            select(ChatMessage)
            .order_by(ChatMessage.created_at.desc())
            .limit(limit)
        ).all()
        
        result = []
        for msg in reversed(messages):
            rule_data = None
            if msg.rule_id:
                rule = session.get(AutomationRule, msg.rule_id)
                if rule:
                    rule_data = {
                        "name": rule.name,
                        "nodes": json.loads(rule.nodes_json),
                        "command": rule.command,
                        "extraction_hint": rule.extraction_hint,
                        "condition": rule.condition,
                        "action": rule.action
                    }
            
            result.append({
                "id": msg.id,
                "role": msg.role,
                "content": msg.content,
                "rule_id": msg.rule_id,
                "rule": rule_data,
                "created_at": msg.created_at.isoformat()
            })
            
        return result


@router.get("/rules")
async def get_rules():
    """Get all rules"""
    with get_sync_session() as session:
        rules = session.exec(select(AutomationRule)).all()
        return [
            {
                "id": rule.id,
                "name": rule.name,
                "command": rule.command,
                "condition": rule.condition,
                "action": rule.action,
                "status": rule.status,
                "schedule": rule.schedule_human,
                "last_executed": rule.last_executed_at.isoformat() if rule.last_executed_at else None
            }
            for rule in rules
        ]


@router.get("/executions")
async def get_executions(limit: int = 20, offset: int = 0):
    """Get recent executions for historical grid"""
    with get_sync_session() as session:
        executions = session.exec(
            select(RuleExecution)
            .order_by(RuleExecution.started_at.desc())
            .offset(offset)
            .limit(limit)
        ).all()
        
        # Get total count for pagination
        # Note: SQLModel count is a bit verbose, we'll skip strict count for now or add if needed
        # Simpler just to return list and client handles "next" logic based on empty list
        
        result = []
        for exe in executions:
            rule = session.get(AutomationRule, exe.rule_id)
            result.append({
                "id": exe.id,
                "rule_name": rule.name if rule else "Unknown Rule",
                "status": exe.status,
                "started_at": exe.started_at.isoformat(),
                "completed_at": exe.completed_at.isoformat() if exe.completed_at else None,
                "condition_met": exe.condition_met,
                "parsed_value": exe.parsed_value
            })
        return result


@router.get("/incidents")
async def get_incidents():
    """Get all incidents"""
    with get_sync_session() as session:
        incidents = session.exec(
            select(Incident).order_by(Incident.created_at.desc())
        ).all()
        return [
            {
                "id": inc.id,
                "rule_id": inc.rule_id,
                "severity": inc.severity,
                "node": inc.node,
                "message": inc.message,
                "metric_value": inc.metric_value,
                "threshold": inc.threshold,
                "acknowledged": inc.acknowledged,
                "resolved": inc.resolved,
                "created_at": inc.created_at.isoformat()
            }
            for inc in incidents
        ]


@router.post("/nodes")
async def add_known_node(
    name: str, 
    hostname: str, 
    username: str = "ubuntu",
    port: int = 22,
    auth_type: str = "private_key",
    credential: str = None
):
    """Add a known node for auto-resolution"""
    from crypto import encrypt_credential
    
    with get_sync_session() as session:
        node = KnownNode(
            name=name,
            hostname=hostname,
            username=username,
            port=port,
            auth_type=auth_type,
            credential_encrypted=encrypt_credential(credential) if credential else None
        )
        session.add(node)
        session.commit()
        session.refresh(node)
        
        return {"id": node.id, "name": node.name, "hostname": node.hostname}
